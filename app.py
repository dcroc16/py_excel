import openpyxl as xl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
ft = Font(name='Calibri',
                 size=12,
                bold=True,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='FF000000')

fill = PatternFill(fill_type=None,
                 start_color='FFFFFFFF',
                 end_color='FF000000')
border = Border(left=Side(border_style=None,
                           color='FF000000'),
                 right=Side(border_style=None,
                            color='FF000000'),
                 top=Side(border_style=None,
                          color='FF000000'),
                 bottom=Side(border_style=None,
                             color='FF000000'),
                 diagonal=Side(border_style=None,
                               color='FF000000'),
                 diagonal_direction=0,
                 outline=Side(border_style=None,
                              color='FF000000'),
                 vertical=Side(border_style=None,
                               color='FF000000'),
                 horizontal=Side(border_style=None,
                                color='FF000000')
                )
alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=False,
                     shrink_to_fit=False,
                     indent=0)
number_format = 'General'
protection = Protection(locked=True,
                         hidden=False)

print("hello world")

wb = xl.load_workbook("example.xlsx")

print(type(wb))
print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name("Sheet1")
print(sheet["A1"].value)

i = 1

while sheet.cell(row=i, column=1).value != None:
    print(sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=2).value)
    sheet.cell(row=i, column=3).value = "=5*" + str(i)
    sheet.cell(row=i, column=3).font = ft
    i += 1


print(sheet.max_row, sheet.max_column)


sheet["A25"].value = "1025"

wb.save("example.xlsx")

